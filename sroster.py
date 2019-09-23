import openpyxl
from openpyxl.styles import Font
import csv
import sys
import os


teacherData = {}
studentData = {}
wb = openpyxl.Workbook()

def extractData(infile):
    with open(infile) as csv_file:

        csv_reader = csv.reader(csv_file)

        
        for data in csv_reader:
            if data and data[0].isdigit():
                studentid = data[0]
                student_last = data[1]
                student_first = data[2]
                term_start = data[4]
                term_end = data[5]
                teacher = data[6]
                period_start = data[7]
                grade = data[8]
                room = data[11]
            
                ### JOIN NAMES FOR EMAIL and CLEANING UP SPEACIAL CHARS ###
                email_first = ''.join(e for e in student_first if e.isalnum())
                email_last = ''.join(e for e in student_last if e.isalnum())

                username = email_first + '.' + email_last
                student_email = email_first + '.' + email_last + "@seariders.k12.hi.us"
                student_pass = student_first[-2:] + studentid[-6:]

                ### STUDENT DICTIONARY FOR STUDENT SHEETS ###
                class_dict = {'teacher' : teacher, 'room' : room, 'period_start': period_start, 'term_start': term_start, 'term_end': term_end}
                
                if studentid not in studentData:
                    classes = []
                    classes.append(class_dict)
                    studentData.setdefault(studentid, {'lastname' : student_last, 'firstname' : student_first, 'grade' : grade , 'email' : student_email.lower(), 'password' : student_pass, 'classes' : classes, 'username' : username })
                else:
                    classes = studentData[studentid]['classes']
                    classes.append(class_dict)
                    studentData[studentid]['classes'] = classes

                ### TEACHER DICTIONARY FOR TEACHER SHEETS ###
                if teacher is '':
                    teacher = 'Activity'
            
                teacherData.setdefault(teacher, {})
                teacherData[teacher].setdefault(studentid, {'lastname' : student_last, 'firstname' : student_first, 'grade' : grade , 'email' : student_email.lower(), 'password' : student_pass, 'period_start' : period_start, 'room': room, 'term_start': term_start, 'term_end' : term_end})
                
    csv_file.close()

def StudentSheet(verbose):
    ########################### "ALL STUDENTS" ##################################
    if verbose:
        print("\nProccessing All Students.... ")

    sheetname1 = "All Students"
    wb.create_sheet(title=sheetname1)
    sheet = wb[sheetname1]

    sheet['A1'] = 'Student Number'
    sheet['B1'] = 'Last Name'
    sheet['C1'] = 'First Name'
    sheet['D1'] = 'Grade'
    sheet['E1'] = 'Seariders Gmail'
    sheet['F1'] = 'Password'

    sheet.freeze_panes = "A2"
    sheet.print_title_rows='1:1'

    rowNum = 2

    for studentid in studentData:

        col1 = sheet.cell(row=rowNum, column=1)
        col2 = sheet.cell(row=rowNum, column=2)
        col3 = sheet.cell(row=rowNum, column=3)
        col4 = sheet.cell(row=rowNum, column=4)
        col5 = sheet.cell(row=rowNum, column=5)
        col6 = sheet.cell(row=rowNum, column=6)
        
        firstname = studentData[studentid]['firstname']
        lastname = studentData[studentid]['lastname']
        grade = studentData[studentid]['grade']
        email = studentData[studentid]['email']
        password = studentData[studentid]['password']

        col1.value = studentid
        col2.value = lastname
        col3.value = firstname
        col4.value = grade
        col5.value = email
        col6.value = password

        rowNum += 1 

def ScheduleSheet(verbose):
    ############################ "SCHEDULES" ##################################
    if verbose:
        print("\nProccessing Schedule Sheet.... ")

    sheetname2 = "Schedules"
    wb.create_sheet(title=sheetname2)
    sheet = wb[sheetname2]
    
    sheet['A1'] = 'Student Number'
    sheet['B1'] = 'Last Name'
    sheet['C1'] = 'First Name'
    sheet['D1'] = 'Grade'
    sheet['E1'] = 'Teacher'
    sheet['F1'] = 'Room Name'
    sheet['G1'] = 'Period Start'
    sheet['H1'] = 'Seariders Gmail'
    sheet['I1'] = 'Password'
    sheet['J1'] = 'Term Start'
    sheet['K1'] = 'Term End'
    sheet['L1'] = 'sort(Term Start)'

    sheet.freeze_panes = "A2"
    sheet.print_title_rows='1:1'

    rowNum = 2

    for studentid in studentData:

        classlist = studentData[studentid]['classes']

        for studentclass in classlist:
            col1 = sheet.cell(row=rowNum, column=1)
            col2 = sheet.cell(row=rowNum, column=2)
            col3 = sheet.cell(row=rowNum, column=3)
            col4 = sheet.cell(row=rowNum, column=4)
            col5 = sheet.cell(row=rowNum, column=5)
            col6 = sheet.cell(row=rowNum, column=6)
            col7 = sheet.cell(row=rowNum, column=7)
            col8 = sheet.cell(row=rowNum, column=8)
            col9 = sheet.cell(row=rowNum, column=9)
            col10 = sheet.cell(row=rowNum, column=10)
            col11 = sheet.cell(row=rowNum, column=11)
            col12 = sheet.cell(row=rowNum, column=12)
        
            firstname = studentData[studentid]['firstname']
            lastname = studentData[studentid]['lastname']
            grade = studentData[studentid]['grade']
            email = studentData[studentid]['email']
            teacher = studentclass['teacher']
            room = studentclass['room']
            period_start = studentclass['period_start']
            password = studentData[studentid]['password']
            term_start = studentclass['term_start']
            term_end = studentclass['term_end']

            col1.value = studentid
            col2.value = lastname
            col3.value = firstname
            col4.value = grade
            col5.value = teacher
            col6.value = room
            col7.value = period_start
            col8.value = email
            col9.value = password
            col10.value = term_start
            col11.value = term_end
            col12.value = term_start[1]

            rowNum += 1

def TeacherSheets(verbose):
    ############################  TEACHERS ###################################
    if verbose:
        print("\nProcessing Teacher Sheets...")

    for teacher in sorted (teacherData.keys()):

        wb.create_sheet(title=teacher)

        if verbose:
            print("\tCreating Sheet for [" + teacher + "]...")

        sheet = wb[teacher]
        
        sheet['A1'] = 'Student Number'
        sheet['B1'] = 'Last Name'
        sheet['C1'] = 'First Name'
        sheet['D1'] = 'Grade'
        sheet['E1'] = 'Teacher'
        sheet['F1'] = 'Room Name'
        sheet['G1'] = 'Period Start'
        sheet['H1'] = 'Seariders Gmail'
        sheet['I1'] = 'Password'
        sheet['J1'] = 'Term Start'
        sheet['K1'] = 'Term End'
        sheet['L1'] = 'sort(Term Start)'


        sheet.freeze_panes = "A2"
        sheet.print_title_rows='1:1'

        rowNum = 2

        for studentid in teacherData[teacher]:

            student = teacherData[teacher][studentid]

            col1 = sheet.cell(row=rowNum, column=1)
            col2 = sheet.cell(row=rowNum, column=2)
            col3 = sheet.cell(row=rowNum, column=3)
            col4 = sheet.cell(row=rowNum, column=4)
            col5 = sheet.cell(row=rowNum, column=5)
            col6 = sheet.cell(row=rowNum, column=6)
            col7 = sheet.cell(row=rowNum, column=7)
            col8 = sheet.cell(row=rowNum, column=8)
            col9 = sheet.cell(row=rowNum, column=9)
            col10 = sheet.cell(row=rowNum, column=10)
            col11 = sheet.cell(row=rowNum, column=11)
            col12 = sheet.cell(row=rowNum, column=12)

            col1.value = studentid
            col2.value = student['lastname']
            col3.value = student['firstname']
            col4.value = student['grade']
            col5.value = student['period_start']
            col6.value = student['room']
            col7.value = student['period_start']
            col8.value = student['email']
            col9.value = student['password']
            col10.value = student['term_start']
            col11.value = student['term_end']
            col12.value = student['term_start'][1]

            rowNum += 1


def StyleSheets(verbose):
    ############################  Styling ###################################
    if verbose:
        print("\nAlmost there! Making it pretty ...")

    font = Font(name='Calibri',
            size=12,
            bold=False,
            italic=False,
            vertAlign=None,
            underline='none',
            strike=False,
            color='FF000000')

    for worksheet in wb:

        for col in worksheet.columns:
            max_length = 0
            column = col[0].column_letter # Get the column name
        # Since Openpyxl 2.6, the column name is  ".column_letter" as .column became the column number (1-based) 
            for cell in col:
                try: # Necessary to avoid error on empty cells
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                    cell.font = font
                except:
                    pass

            adjusted_width = (max_length + 2) * 1.2
            worksheet.column_dimensions[column].width = adjusted_width

def papercutCardID(outputfile, verbose):

    if verbose:
        print("\nCreating PaperCut import file")

    output = os.path.splitext(outputfile)[0]
    paperfile = open( 'papercut-'+output+'.txt', 'w')


    for studentID in studentData:
        username = studentData[studentID]['username']
        paperfile.write(username + '\t' + studentID + '\n')

            

def main():

    # include standard modules
    import argparse

    # define the program description
    text = 'This program parses infinite campus csv extract file and sorts students under their assigned\
            teachers. Optional: Outputs a import file for PaperCut ID to update papercut users with their IDs.'

    # initiate the parser with a description
    parser = argparse.ArgumentParser(description = text)

    #Required Positional
    parser.add_argument("input", help="Infinite Campus csv file")
    parser.add_argument("output", help="XLSX output file")

    #Optional Args
    parser.add_argument("--paper", "-p", help="Create PaperCut import file to update users with ID numbers",
                        action="store_true")
    
    parser.add_argument("--verbose", "-v", help="See prints",
                        action="store_true")

    args = parser.parse_args()

    try:
        extractData(args.input)
    except:
        print('\n\nERROR: Cant find file ' + args.input + '.')
        print('EXITING PROGRAM')
        sys.exit()
        
    print("Running Script...")
    StudentSheet(args.verbose)
    ScheduleSheet(args.verbose)
    TeacherSheets(args.verbose)
    StyleSheets(args.verbose)

    if args.paper:
        papercutCardID(args.output, args.verbose)
    
    try:
        wb.remove(wb['Sheet'])
    except:
        pass

    wb.save(args.output)

    print("\nDone!")

 
if __name__ == "__main__":
    main()